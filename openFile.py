from docx import Document
from openpyxl import Workbook
import os
import jieba
with open('./source/dic.txt', 'r') as f:   
    result = f.read()       

data_dict = eval(result)

level_count = {"一级": 0,"二级": 0,"三级": 0,"四级": 0,"五级": 0,"六级": 0,"七-九级": 0}
floder_path = "./source"

# 创建一个新的Excel工作簿
workbook = Workbook()
# 创建一个工作表
worksheet = workbook.active
# 设置列名

row_index = 2
worksheet.cell(row=1, column=1, value="文件名+生词频率")

docx_files = [file for file in os.listdir(floder_path) if file.endswith(".docx")]
for file_name in docx_files:
    level_copy = {}
    level_copy = level_count.copy()
    data_dict_copy = data_dict.copy()
    result_dic = {}
    file_path = os.path.join(floder_path, file_name)
    doc = Document(file_path)
    print(file_name)
    print("==============")
    for paragraph in doc.paragraphs:
        text = paragraph.text
        
        seg_list = jieba.cut(text, cut_all=False)
    
        for k in seg_list:
            if k in data_dict_copy:
                level = data_dict_copy[k]
                level_copy[level] += 1
                print(k + ":" + data_dict_copy[k])
                result_dic[k] = data_dict_copy[k]

                del data_dict_copy[k]
            else:
                continue
        # 将文件名写入Excel第一列
        string_res = file_name + ": " +str(level_copy)
    worksheet.cell(row=row_index, column=1, value=string_res)
                # 将字典的键值对写入Excel第二列和第三列
    for i, (key, value) in enumerate(result_dic.items(), start=2):
        worksheet.cell(row=row_index, column=i, value=key)
        worksheet.cell(row=row_index + 1, column=i, value=value)

    # 更新行索引
    row_index += 2

    
workbook.save("./source/output.xlsx")


